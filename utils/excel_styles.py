import io

from copy import copy
import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from utils.data_helpers import converter_valor_inteligente

def copiar_estilo_completo(origem, destino):
    if origem.has_style:
        destino.font = copy(origem.font)
        destino.border = copy(origem.border)
        destino.fill = copy(origem.fill)
        destino.number_format = copy(origem.number_format)
        destino.protection = copy(origem.protection)
        destino.alignment = copy(origem.alignment)

def deduplicar_colunas(colunas):
    vistos = {}
    novas_colunas = []
    for col in colunas:
        col_str = str(col).strip()
        if col_str in vistos:
            vistos[col_str] += 1
            novas_colunas.append(f"{col_str} ({vistos[col_str]})")
        else:
            vistos[col_str] = 1
            novas_colunas.append(col_str)
    return novas_colunas

def formatar_datas_dataframe(df_input):
    df_out = df_input.copy()
    for col in df_out.columns:
        if pd.api.types.is_datetime64_any_dtype(df_out[col]):
            df_out[col] = df_out[col].dt.strftime('%d/%m/%Y').fillna('')
        else:
            df_out[col] = df_out[col].apply(
                lambda v: "" if pd.isna(v) else (
                    v.strftime('%d/%m/%Y') if isinstance(v, (datetime.datetime, datetime.date, pd.Timestamp))
                    else (str(v).split(' ')[0] if isinstance(v, str) and (' 00:00:00' in str(v) or 'T00:00:00' in str(v)) else v)
                )
            )
    return df_out

def obter_nome_coluna_por_letra(df, colunas_disponiveis, letra):
    mapa_letras = {
        'A': 0, 'B': 1, 'C': 2, 'D': 3, 'E': 4, 'F': 5, 'G': 6, 'H': 7,
        'I': 8, 'J': 9, 'K': 10, 'L': 11, 'M': 12, 'N': 13, 'O': 14,
        'P': 15, 'Q': 16, 'R': 17, 'S': 18, 'T': 19, 'U': 20, 'V': 21,
        'W': 22, 'X': 23, 'Y': 24, 'Z': 25
    }
    idx = mapa_letras.get(letra.upper())
    if idx is not None and idx < len(colunas_disponiveis):
        return colunas_disponiveis[idx]
    return None

def gerar_arquivo_atualizado_bytes(source_input, header, fila, df_original, sheet_name=None):
    wb = load_workbook(io.BytesIO(source_input) if isinstance(source_input, bytes) else source_input)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
    for mod in fila:
        col_target = mod['coluna']
        valor_convertido = converter_valor_inteligente(mod['novo_valor'], df_original[col_target].dtype)
        for idx in mod['indices']:
            excel_row = idx + header + 1
            ws.cell(row=excel_row, column=df_original.columns.get_loc(col_target) + 1, value=valor_convertido)

            if col_target.strip().upper() in ["SAIDA", "SAÍDA"]:
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=excel_row, column=col_idx)
                    current_font = cell.font
                    if current_font:
                        cell.font = Font(
                            name=current_font.name,
                            size=current_font.size,
                            bold=current_font.bold,
                            italic=current_font.italic,
                            strike=current_font.strike,
                            underline=current_font.underline,
                            color="FF0000"
                        )
                    else:
                        cell.font = Font(color="FF0000")

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
